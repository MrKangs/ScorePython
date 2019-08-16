import openpyxl

exf = openpyxl.load_workbook("D:\\Python\\ScorePython\\num2.xlsx")
aws = exf.active

print("====================================================")
print("정수\t대수\t학번\t\t성명\t총점\t평균")
print("====================================================")

for i in aws.rows:
    index = i[0].row
    nums1 = i[0].value
    nums2 = i[1].value
    idd = i[2].value
    name = i[3].value
    total = nums1+nums2
    avg = total/2
    aws.cell(row=index, column = 6).value = total    
    aws.cell(row=index, column = 7).value = avg    
    print("{}\t{}\t{}\t{}\t{}\t{:.1f}".format(nums1, nums2, idd, name, total, avg))

exf.save("D:\\Python\\ScorePython\\num3.xlsx")
exf.close()